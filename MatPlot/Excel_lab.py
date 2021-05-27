import openpyxl
import matplotlib.pyplot as plt
import numpy as np

wb = openpyxl.load_workbook('users.xlsx')
sheet = wb.get_sheet_by_name("data")
states = [] #the states are the x values
gender_array = [] #number of users for this gender in x state
storage = {} #used to store the counts for each state of each gender

counter= 0 
for i in range(2,sheet.max_row+1,1):
	if (sheet.cell(row= i, column =6).value) in states:
		pass
	else:
		states.append((sheet.cell(row= i, column =6).value))
		counter+=1

		

for state in states:
	female = 0
	male = 0
	for i in range(2,sheet.max_row+1,1):
		if (sheet.cell(row= i, column =6).value) == state and (sheet.cell(row= i, column =5).value) == 'Female':
			female += 1
		elif (sheet.cell(row= i, column =6).value) == state and (sheet.cell(row= i, column =5).value) == 'Male':
			male += 1
	storage[state] = {'Male':male,'Female':female}

for each in states:
	print('{} {}'.format(each, storage[each]))
print("{}".format(storage[states[1]]))
print("{}".format(storage.values())) 

np.random.seed(19680801)

fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')

colors = ['r', 'b']
yticks = [1, 0]

for c, k in zip(colors, yticks):
    # Generate the random data for the y=k 'layer'.
	xs = range(48)
	males = []
	females = []

	for i in xs:
		if (k == 1):
			females.append(storage[states[i]]['Female'])
			ys = females
		else:
			males.append(storage[states[i]]['Male'])
			ys = males

	
	cs = [c] * len(xs)
	#cs[0] = 'c'

	# Plot the bar graph given by xs and ys on the plane y=k with 80% opacity.
	ax.bar(xs, ys, zs=k, zdir='y', color=cs, alpha=0.8)

ax.set_xlabel('States')
ax.set_ylabel('Genders')
ax.set_zlabel('Number of Users')

# On the y axis let's only label the discrete values that we have data for.
ax.set_yticks(yticks)

plt.show()



